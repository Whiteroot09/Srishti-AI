import speech_recognition as sr
import pyttsx3
import datetime
from openpyxl import load_workbook

def online(text):
    engine = pyttsx3.init()
    voices = engine.getProperty('voices')
    engine.setProperty('voice', voices[3].id)
    engine.setProperty('rate', 180-200)
    print(f"SRIshti Said : {text}")
    engine.say(text)
    engine.runAndWait()


def takeCommand():
    r = sr.Recognizer()
    with sr.Microphone(device_index=0) as source:
        print('listening...')
        audio = r.listen(source, 0, 4)

    try:
        print("recognizing...")
        query = r.recognize_google(audio, language="en-in")
        print(f"User Said : {query}")
    except Exception as e:
        print(e)
        return ""

    query = str(query)
    return query.lower()


def wish_us():
    hour = int(datetime.datetime.now().hour)
    if hour>=0 and hour<12:
        speak("Good morning sir")
        # print(hour)

    elif hour>=12 and hour <= 16:
        speak("Good afternoon sir")
        # print(hour)

    else:
        speak('Good evening sir')
        # print(hour)

def database(quest, ans):
    now = datetime.datetime.now()
    date = now.strftime("%d/%m/%y")
    time = now.strftime("%H:%M:%S")
    wb = load_workbook("D:\\FInalAI\\database.xlsx")
    ws = wb['database_main']
    ws.append([date, time, quest, ans])
    wb.save("D:\\FInalAI\\database.xlsx")

def database_main(quest,ans):
    now = datetime.datetime.now()
    date = now.strftime("%d/%m/%y")
    time = now.strftime("%H:%M:%S")
    wb = load_workbook("D:\\FInalAI\\database.xlsx")
    ws = wb['database_setup']
    ws.append([date, time, quest, ans])
    wb.save("D:\\FInalAI\\database.xlsx")

# takeCommand()